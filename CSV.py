import csv
import openpyxl

class CSV:
  def DictRead(self,path,**kwargs):
    with open(path,"r") as f:
      reader = csv.DictReader(f,delimiter=kwargs.get("delimiter",","))
      return list(reader)
  def DictWrite(self,x,path,**kwargs):
    with open(path,"w") as f:
      writer = csv.DictWriter(f,fieldnames=x[0].keys(),delimiter=kwargs.get("delimiter",","))
      writer.writeheader()
      writer.writerows(x)
  def CsvToXlsx(self,path,**kwargs):
    with open(path,"r") as f:
      reader = csv.reader(f,delimiter=kwargs.get("delimiter",","))
      data = list(reader)
      wb = openpyxl.Workbook()
      ws = wb["Sheet"]
      for idx, i in enumerate(data):
        for idx2, j in enumerate(i):
          ws.cell(idx+1,idx2+1).value = j
      wb.save("Workbook.xlsx")
  def XlsxToCsv(self,path,sheet,**kwargs):
    wb = openpyxl.load_workbook(path)
    ws = wb[sheet]
    values = list(ws.values)
    headers = values[0]
    data = values[1:]
    x = list(map(lambda i: dict(zip(headers,i)),data))
    CSV().DictWrite(x,"Csv.csv",delimiter=",")
